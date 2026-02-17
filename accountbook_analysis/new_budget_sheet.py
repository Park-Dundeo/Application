"""
새 예산안 시트 생성 스크립트
Google Sheets API로 직접 생성하거나, 로컬에서 XLSX로 미리보기 가능

사용법:
  python3 accountbook_analysis/new_budget_sheet.py          # XLSX 미리보기 생성
  python3 accountbook_analysis/new_budget_sheet.py --apply   # Google Sheets에 적용
"""

import argparse
import os
import sys

# ── 새 예산안 데이터 정의 ──────────────────────────────────────────────

BUDGET_PERIOD = "26.7~27.6"
ANNUAL_INCOME = 60_000_000
MONTHLY_BASE_SALARY = 3_500_000
BONUS_ANNUAL = 17_500_000

# 층별 예산 항목: (구분, 프로젝트, 항목명, 월예산, 연예산, 비고)
BUDGET_ITEMS = [
    # ── 1층: 고정 지출 ──
    ("고정", "운영", "자동차 할부",  610_000,  7_320_000, "할부금 고정"),
    ("고정", "운영", "보험",        160_000,  1_920_000, "실적 153K"),
    ("고정", "운영", "통신",        130_000,  1_560_000, "실적 122K"),
    ("고정", "운영", "커플통장",     200_000,  2_400_000, "약정"),
    ("고정", "효도", "어머니 적금",  300_000,  3_600_000, "약정"),
    ("고정", "운영", "AI",          15_000,    180_000, "구독"),

    # ── 2층: 생활 지출 ──
    ("생활", "생활", "식비",        800_000,  9_600_000, "실적972K→800K 절감목표"),
    ("생활", "생활", "교통",         80_000,    960_000, "실적 77K"),
    ("생활", "생활", "카페/간식",    130_000,  1_560_000, "실적 115K"),
    ("생활", "생활", "자동차 유지",  200_000,  2_400_000, "주유/톨비/주차 월상"),
    ("생활", "생활", "생활용품",     100_000,  1_200_000, "생활잡화"),
    ("생활", "생활", "온라인쇼핑",   100_000,  1_200_000, "실적 109K→100K"),
    ("생활", "생활", "뷰티/미용",     50_000,    600_000, "실적 75K→50K"),
    ("생활", "개발", "도서",         20_000,    240_000, "실적 10K"),

    # ── 3층: 프로젝트 지출 ──
    ("프로젝트", "효도",     "어머니 용돈",   500_000,  6_000_000, "실적 513K"),
    ("프로젝트", "효도",     "어머니 선물",   150_000,  1_800_000, "월상 소형 선물"),
    ("프로젝트", "선교",     "선교 헌금",     120_000,  1_440_000, "실적 119K"),
    ("프로젝트", "선교",     "필리핀",         10_000,    120_000, "실적 10K"),
    ("프로젝트", "선교",     "수련회",         15_000,    180_000, "연 1~2회"),
    ("프로젝트", "찬양",     "찬양팀 회비",    15_000,    180_000, "실적 일치"),
    ("프로젝트", "교회/양육", "식사교제",       10_000,    120_000, "실적 7K"),
    ("프로젝트", "ACS",      "모임비",          0,            0, "현재 미발생"),
    ("프로젝트", "관계",     "선물",          100_000,  1_200_000, "월상 소형 선물"),
    ("프로젝트", "관계",     "동기 계모임",    15_000,    180_000, "실적 일치"),
    ("프로젝트", "관계",     "파트 회비",      20_000,    240_000, "격월 20K"),
    ("프로젝트", "경조사",   "경조비",         60_000,    720_000, "실적 117K, 보수적"),

    # ── 4층: 비정기/연간 (적립) ──
    ("비정기", "건강",   "의료/건강",        120_000,  1_440_000, "병원 실적 231K, 변동대비"),
    ("비정기", "운영",   "자동차 유지(연간)", 150_000, 1_800_000, "보험갱신/정비/수리"),
    ("비정기", "여행",   "여행",             200_000,  2_400_000, "여수+일본+속초 등"),
    ("비정기", "관계",   "대형선물",         100_000,  1_200_000, "비정기 대형 선물"),
    ("비정기", "효도",   "효도 이벤트",      100_000,  1_200_000, "어머니 여행/큰선물"),
    ("비정기", "생활",   "패션/쇼핑",         80_000,    960_000, "실적 219K→절감"),
    ("비정기", "생활",   "문화/여가",         80_000,    960_000, "실적 413K→대폭절감"),
    ("비정기", "예비",   "예비비",           100_000,  1_200_000, "순수 돌발지출"),
]

TIER_NAMES = {
    "고정": "1층: 고정 지출",
    "생활": "2층: 생활 지출",
    "프로젝트": "3층: 프로젝트 지출",
    "비정기": "4층: 비정기/연간 적립",
}

TIER_COLORS = {
    "고정":    {"bg": "#FFF2CC", "header": "#F1C232"},  # 노랑
    "생활":    {"bg": "#D9EAD3", "header": "#6AA84F"},  # 초록
    "프로젝트": {"bg": "#CFE2F3", "header": "#3D85C6"},  # 파랑
    "비정기":  {"bg": "#F4CCCC", "header": "#CC0000"},  # 분홍
}


def create_xlsx_preview():
    """로컬 XLSX 파일로 미리보기 생성"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{BUDGET_PERIOD} 예산안 (신규)"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ── 열 너비 설정 ──
    col_widths = {'A': 10, 'B': 12, 'C': 18, 'D': 12, 'E': 14,
                  'F': 14, 'G': 14, 'H': 14, 'I': 30}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── 대시보드 영역 (Row 1~5) ──
    ws['A1'] = f"{BUDGET_PERIOD} 예산안"
    ws['A1'].font = Font(bold=True, size=14)

    dashboard_items = [
        ('A2', '월 기본급'),  ('B2', MONTHLY_BASE_SALARY),
        ('C2', '연간 수입'),  ('D2', ANNUAL_INCOME),
        ('E2', '상여금/년'),  ('F2', BONUS_ANNUAL),
        ('A3', '경과 월수'),  ('B3', '=MONTH(NOW())-MONTH(DATE(2026,7,1))+1'),
        ('C3', '전월 월급'),
        ('D3', "=SUMIFS('가계부 내역'!G:G,'가계부 내역'!A:A,\">=\"&EOMONTH(TODAY(),-2)+1,'가계부 내역'!A:A,\"<=\"&EOMONTH(TODAY(),-1),'가계부 내역'!F:F,\"급여\")"),
        ('E3', '본월 지출'),
        ('F3', "=SUMIFS('가계부 내역'!G:G,'가계부 내역'!A:A,\">=\"&EOMONTH(TODAY(),-1)+1,'가계부 내역'!A:A,\"<=\"&EOMONTH(TODAY(),0),'가계부 내역'!C:C,\"지출\")"),
    ]
    for cell_ref, value in dashboard_items:
        ws[cell_ref] = value
        ws[cell_ref].border = thin_border

    # Row 4: 소비 합계 대시보드
    ws['A4'] = '월 소비예산'
    ws['B4'] = f'=SUM(D{7}:D{7 + len(BUDGET_ITEMS) + 10})'  # 나중에 정확한 범위로
    ws['C4'] = '연 소비예산'
    ws['D4'] = f'=SUM(E{7}:E{7 + len(BUDGET_ITEMS) + 10})'
    ws['E4'] = '월 잔여'
    ws['F4'] = '=B2-B4'
    ws['G4'] = '연 잔여'
    ws['H4'] = '=D2-D4'

    for row in ws.iter_rows(min_row=2, max_row=4, max_col=8):
        for cell in row:
            cell.border = thin_border
            if cell.column in (2, 4, 6, 8):
                cell.number_format = '#,##0'

    # ── 헤더 행 (Row 6) ──
    headers = ['구분', '프로젝트', '항목', '월 예산', '연 예산',
               '실적', '남은 예산', '여분/부족', '비고']
    header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # ── 데이터 행 (Row 7~) ──
    current_row = 7
    current_tier = None
    tier_start_rows = {}
    tier_end_rows = {}

    for tier, project, item_name, monthly, annual, note in BUDGET_ITEMS:
        # 층 구분이 바뀔 때 소계 헤더 삽입
        if tier != current_tier:
            if current_tier is not None:
                # 이전 층 소계 행
                tier_end_rows[current_tier] = current_row - 1
                subtotal_fill = PatternFill(
                    start_color=TIER_COLORS[current_tier]['header'].replace('#', ''),
                    end_color=TIER_COLORS[current_tier]['header'].replace('#', ''),
                    fill_type='solid'
                )
                ws.cell(row=current_row, column=1, value='').fill = subtotal_fill
                ws.cell(row=current_row, column=2, value='').fill = subtotal_fill
                ws.cell(row=current_row, column=3, value=f'{TIER_NAMES[current_tier]} 소계').fill = subtotal_fill
                ws.cell(row=current_row, column=3).font = Font(bold=True, color='FFFFFF')
                start = tier_start_rows[current_tier]
                ws.cell(row=current_row, column=4,
                        value=f'=SUM(D{start}:D{current_row - 1})').fill = subtotal_fill
                ws.cell(row=current_row, column=4).font = Font(bold=True, color='FFFFFF')
                ws.cell(row=current_row, column=4).number_format = '#,##0'
                ws.cell(row=current_row, column=5,
                        value=f'=SUM(E{start}:E{current_row - 1})').fill = subtotal_fill
                ws.cell(row=current_row, column=5).font = Font(bold=True, color='FFFFFF')
                ws.cell(row=current_row, column=5).number_format = '#,##0'
                ws.cell(row=current_row, column=6,
                        value=f'=SUM(F{start}:F{current_row - 1})').fill = subtotal_fill
                ws.cell(row=current_row, column=6).font = Font(bold=True, color='FFFFFF')
                ws.cell(row=current_row, column=6).number_format = '#,##0'
                ws.cell(row=current_row, column=7,
                        value=f'=SUM(G{start}:G{current_row - 1})').fill = subtotal_fill
                ws.cell(row=current_row, column=7).font = Font(bold=True, color='FFFFFF')
                ws.cell(row=current_row, column=7).number_format = '#,##0'
                ws.cell(row=current_row, column=8,
                        value=f'=SUM(H{start}:H{current_row - 1})').fill = subtotal_fill
                ws.cell(row=current_row, column=8).font = Font(bold=True, color='FFFFFF')
                ws.cell(row=current_row, column=8).number_format = '#,##0'
                for c in range(1, 10):
                    ws.cell(row=current_row, column=c).border = thin_border
                current_row += 1

            current_tier = tier
            tier_start_rows[tier] = current_row

        # 데이터 행
        row_fill = PatternFill(
            start_color=TIER_COLORS[tier]['bg'].replace('#', ''),
            end_color=TIER_COLORS[tier]['bg'].replace('#', ''),
            fill_type='solid'
        )

        row_data = [
            tier,          # A: 구분
            project,       # B: 프로젝트
            item_name,     # C: 항목
            monthly,       # D: 월 예산
            annual,        # E: 연 예산
            # F: 실적 = SUMIF(가계부내역 K열, 항목명, G열)
            f"=SUMIF('가계부 내역'!$K:$K,C{current_row},'가계부 내역'!$G:$G)",
            # G: 남은 예산 = 연예산 + 실적
            f"=E{current_row}+F{current_row}",
            # H: 여분/부족 = 월예산 × 경과월 + 실적
            f"=D{current_row}*$B$3+F{current_row}",
            note,          # I: 비고
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx in (4, 5, 6, 7, 8):
                cell.number_format = '#,##0'

        current_row += 1

    # 마지막 층 소계
    if current_tier:
        tier_end_rows[current_tier] = current_row - 1
        subtotal_fill = PatternFill(
            start_color=TIER_COLORS[current_tier]['header'].replace('#', ''),
            end_color=TIER_COLORS[current_tier]['header'].replace('#', ''),
            fill_type='solid'
        )
        start = tier_start_rows[current_tier]
        ws.cell(row=current_row, column=3, value=f'{TIER_NAMES[current_tier]} 소계').fill = subtotal_fill
        ws.cell(row=current_row, column=3).font = Font(bold=True, color='FFFFFF')
        for c in [4, 5, 6, 7, 8]:
            ws.cell(row=current_row, column=c,
                    value=f'=SUM({get_column_letter(c)}{start}:{get_column_letter(c)}{current_row - 1})')
            ws.cell(row=current_row, column=c).fill = subtotal_fill
            ws.cell(row=current_row, column=c).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=current_row, column=c).number_format = '#,##0'
        for c in range(1, 10):
            ws.cell(row=current_row, column=c).border = thin_border
            if not ws.cell(row=current_row, column=c).fill.start_color.rgb or \
               ws.cell(row=current_row, column=c).fill.start_color.rgb == '00000000':
                ws.cell(row=current_row, column=c).fill = subtotal_fill
        current_row += 1

    # ── 총계 행 ──
    current_row += 1
    total_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    total_font = Font(bold=True, color='FFFFFF', size=11)

    ws.cell(row=current_row, column=3, value='총계')
    for c in range(1, 10):
        ws.cell(row=current_row, column=c).fill = total_fill
        ws.cell(row=current_row, column=c).font = total_font
        ws.cell(row=current_row, column=c).border = thin_border

    # 각 층 소계 행 번호 수집
    subtotal_rows = []
    for tier in ['고정', '생활', '프로젝트', '비정기']:
        if tier in tier_end_rows:
            subtotal_rows.append(tier_end_rows[tier] + 1)

    for c in [4, 5, 6, 7, 8]:
        col_letter = get_column_letter(c)
        formula = '+'.join(f'{col_letter}{r}' for r in subtotal_rows)
        ws.cell(row=current_row, column=c, value=f'={formula}')
        ws.cell(row=current_row, column=c).number_format = '#,##0'

    # 대시보드의 합계 참조 수정
    ws['B4'] = f'=D{current_row}'
    ws['D4'] = f'=E{current_row}'

    # ── 저축/투자 영역 ──
    current_row += 2
    ws.cell(row=current_row, column=1, value='자산 운용 계획').font = Font(bold=True, size=12)
    current_row += 1

    savings_data = [
        ('월 잔여 (기본급-소비)', f'=B2-B4'),
        ('연 잔여', f'=D2-D4'),
        ('상여금', BONUS_ANNUAL),
        ('상여→투자(주식)', 10_000_000),
        ('상여→저축', 5_000_000),
        ('상여→여유분', 2_500_000),
        ('연간 자산축적 목표', f'={current_row}+{current_row+2}'),
    ]

    for label, value in savings_data:
        ws.cell(row=current_row, column=3, value=label)
        ws.cell(row=current_row, column=5, value=value)
        ws.cell(row=current_row, column=5).number_format = '#,##0'
        ws.cell(row=current_row, column=3).border = thin_border
        ws.cell(row=current_row, column=5).border = thin_border
        current_row += 1

    # 저장
    output_path = os.path.join(os.path.dirname(__file__), f'new_budget_{BUDGET_PERIOD}.xlsx')
    wb.save(output_path)
    print(f"미리보기 생성 완료: {output_path}")
    return output_path


def _hex_to_rgb(hex_color):
    """'#RRGGBB' → {'red': 0~1, 'green': 0~1, 'blue': 0~1}"""
    h = hex_color.lstrip('#')
    return {
        'red': int(h[0:2], 16) / 255,
        'green': int(h[2:4], 16) / 255,
        'blue': int(h[4:6], 16) / 255,
    }


def _cell_format_request(sheet_id, row, col, col_end=None, **kwargs):
    """단일/범위 셀 서식 요청 생성 헬퍼"""
    if col_end is None:
        col_end = col + 1
    fields = []
    cell_format = {}

    bg = kwargs.get('bg')
    if bg:
        cell_format['backgroundColor'] = _hex_to_rgb(bg)
        fields.append('userEnteredFormat.backgroundColor')

    bold = kwargs.get('bold')
    font_color = kwargs.get('font_color')
    font_size = kwargs.get('font_size')
    if bold is not None or font_color or font_size:
        tf = {}
        if bold is not None:
            tf['bold'] = bold
        if font_color:
            tf['foregroundColor'] = _hex_to_rgb(font_color)
        if font_size:
            tf['fontSize'] = font_size
        cell_format['textFormat'] = tf
        fields.append('userEnteredFormat.textFormat')

    h_align = kwargs.get('h_align')
    if h_align:
        cell_format['horizontalAlignment'] = h_align
        fields.append('userEnteredFormat.horizontalAlignment')

    num_fmt = kwargs.get('num_fmt')
    if num_fmt:
        cell_format['numberFormat'] = {'type': 'NUMBER', 'pattern': num_fmt}
        fields.append('userEnteredFormat.numberFormat')

    borders = kwargs.get('borders')
    if borders:
        border_style = {'style': 'SOLID', 'color': _hex_to_rgb('#CCCCCC')}
        cell_format['borders'] = {
            'top': border_style, 'bottom': border_style,
            'left': border_style, 'right': border_style,
        }
        fields.append('userEnteredFormat.borders')

    return {
        'repeatCell': {
            'range': {
                'sheetId': sheet_id,
                'startRowIndex': row,
                'endRowIndex': row + 1,
                'startColumnIndex': col,
                'endColumnIndex': col_end,
            },
            'cell': {'userEnteredFormat': cell_format},
            'fields': ','.join(fields),
        }
    }


def _row_format_request(sheet_id, row, col_end=9, **kwargs):
    """행 전체 서식"""
    return _cell_format_request(sheet_id, row, 0, col_end, **kwargs)


def create_google_sheet():
    """Google Sheets API로 새 시트 생성 (기존 스프레드시트에 탭 추가)"""
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

    from app.adapters.google_auth import get_credentials
    from googleapiclient.discovery import build

    creds = get_credentials(scopes=[
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive',
    ])
    service = build('sheets', 'v4', credentials=creds)

    spreadsheet_id = os.environ.get('SPREADSHEET_ID')
    if not spreadsheet_id:
        print("ERROR: SPREADSHEET_ID 환경변수 필요")
        sys.exit(1)

    sheet_title = f"{BUDGET_PERIOD} 예산안 (신규)"

    # 1. 새 시트 추가
    add_sheet_req = {
        'addSheet': {
            'properties': {
                'title': sheet_title,
                'gridProperties': {'rowCount': 60, 'columnCount': 10}
            }
        }
    }

    result = service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={'requests': [add_sheet_req]}
    ).execute()

    sheet_id = result['replies'][0]['addSheet']['properties']['sheetId']
    print(f"시트 생성 완료: {sheet_title} (ID: {sheet_id})")

    # 2. 데이터 구성 + 행 메타 추적
    values = []
    row_meta = []  # (row_index, type, tier_or_none)

    # Row 0: 제목
    values.append([f"{BUDGET_PERIOD} 예산안"])
    row_meta.append((0, 'title', None))

    # Row 1: 대시보드 1
    values.append(['월 기본급', MONTHLY_BASE_SALARY, '연간 수입', ANNUAL_INCOME,
                    '상여금/년', BONUS_ANNUAL])
    row_meta.append((1, 'dashboard', None))

    # Row 2: 대시보드 2
    values.append(['경과 월수', '=MONTH(NOW())-MONTH(DATE(2026,7,1))+1',
                    '전월 월급',
                    "=SUMIFS('가계부 내역'!G:G,'가계부 내역'!A:A,\">=\"&EOMONTH(TODAY(),-2)+1,'가계부 내역'!A:A,\"<=\"&EOMONTH(TODAY(),-1),'가계부 내역'!F:F,\"급여\")",
                    '본월 지출',
                    "=SUMIFS('가계부 내역'!G:G,'가계부 내역'!A:A,\">=\"&EOMONTH(TODAY(),-1)+1,'가계부 내역'!A:A,\"<=\"&EOMONTH(TODAY(),0),'가계부 내역'!C:C,\"지출\")"])
    row_meta.append((2, 'dashboard', None))

    # Row 3: 대시보드 합계 (나중에 수식 설정)
    values.append(['월 소비예산', '', '연 소비예산', '', '월 잔여', '', '연 잔여', ''])
    row_meta.append((3, 'dashboard', None))

    # Row 4: 빈 행
    values.append([])
    row_meta.append((4, 'blank', None))

    # Row 5: 헤더
    values.append(['구분', '프로젝트', '항목', '월 예산', '연 예산',
                    '실적', '남은 예산', '여분/부족', '비고'])
    row_meta.append((5, 'header', None))

    # 데이터 행
    current_tier = None
    tier_start_rows = {}  # 0-indexed
    subtotal_row_indices = []  # 0-indexed

    for tier, project, item_name, monthly, annual, note in BUDGET_ITEMS:
        if tier != current_tier and current_tier is not None:
            # 소계 행
            start = tier_start_rows[current_tier]
            ri = len(values)
            gs_row = ri + 1  # 1-indexed for formulas
            gs_start = start + 1
            gs_end = gs_row - 1
            values.append([
                '', '', f'{TIER_NAMES[current_tier]} 소계',
                f'=SUM(D{gs_start}:D{gs_end})',
                f'=SUM(E{gs_start}:E{gs_end})',
                f'=SUM(F{gs_start}:F{gs_end})',
                f'=SUM(G{gs_start}:G{gs_end})',
                f'=SUM(H{gs_start}:H{gs_end})',
                ''
            ])
            row_meta.append((ri, 'subtotal', current_tier))
            subtotal_row_indices.append(ri)

        if tier != current_tier:
            current_tier = tier
            tier_start_rows[tier] = len(values)

        ri = len(values)
        gs_row = ri + 1
        values.append([
            tier, project, item_name, monthly, annual,
            f"=SUMIF('가계부 내역'!$K:$K,C{gs_row},'가계부 내역'!$G:$G)",
            f"=E{gs_row}+F{gs_row}",
            f"=D{gs_row}*$B$3+F{gs_row}",
            note
        ])
        row_meta.append((ri, 'data', tier))

    # 마지막 층 소계
    if current_tier:
        start = tier_start_rows[current_tier]
        ri = len(values)
        gs_row = ri + 1
        gs_start = start + 1
        gs_end = gs_row - 1
        values.append([
            '', '', f'{TIER_NAMES[current_tier]} 소계',
            f'=SUM(D{gs_start}:D{gs_end})',
            f'=SUM(E{gs_start}:E{gs_end})',
            f'=SUM(F{gs_start}:F{gs_end})',
            f'=SUM(G{gs_start}:G{gs_end})',
            f'=SUM(H{gs_start}:H{gs_end})',
            ''
        ])
        row_meta.append((ri, 'subtotal', current_tier))
        subtotal_row_indices.append(ri)

    # 빈 행 + 총계
    values.append([])
    row_meta.append((len(values) - 1, 'blank', None))

    total_ri = len(values)
    gs_total = total_ri + 1
    subtotal_refs_d = '+'.join(f'D{si + 1}' for si in subtotal_row_indices)
    subtotal_refs_e = '+'.join(f'E{si + 1}' for si in subtotal_row_indices)
    subtotal_refs_f = '+'.join(f'F{si + 1}' for si in subtotal_row_indices)
    subtotal_refs_g = '+'.join(f'G{si + 1}' for si in subtotal_row_indices)
    subtotal_refs_h = '+'.join(f'H{si + 1}' for si in subtotal_row_indices)
    values.append([
        '', '', '총계',
        f'={subtotal_refs_d}', f'={subtotal_refs_e}',
        f'={subtotal_refs_f}', f'={subtotal_refs_g}',
        f'={subtotal_refs_h}', ''
    ])
    row_meta.append((total_ri, 'total', None))

    # Row 3 대시보드 합계 수식 설정
    values[3] = [
        '월 소비예산', f'=D{gs_total}',
        '연 소비예산', f'=E{gs_total}',
        '월 잔여', f'=B2-B4',
        '연 잔여', f'=D2-D4',
    ]

    # 3. 데이터 입력
    range_name = f"'{sheet_title}'!A1"
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueInputOption='USER_ENTERED',
        body={'values': values}
    ).execute()

    print(f"데이터 입력 완료: {len(values)}행")

    # 4. 서식 적용
    fmt_requests = []

    # 열 너비 (pixels)
    col_widths_px = [80, 100, 160, 100, 120, 120, 120, 120, 250]
    for i, px in enumerate(col_widths_px):
        fmt_requests.append({
            'updateDimensionProperties': {
                'range': {
                    'sheetId': sheet_id,
                    'dimension': 'COLUMNS',
                    'startIndex': i,
                    'endIndex': i + 1,
                },
                'properties': {'pixelSize': px},
                'fields': 'pixelSize',
            }
        })

    # 제목 행 (Row 0)
    fmt_requests.append(_row_format_request(
        sheet_id, 0, bold=True, font_size=14))

    # 대시보드 (Row 1~3) - 테두리 + 숫자포맷
    for r in range(1, 4):
        fmt_requests.append(_row_format_request(
            sheet_id, r, col_end=8, borders=True))
        # 짝수 열(B,D,F,H = index 1,3,5,7)에 숫자 포맷
        for c in [1, 3, 5, 7]:
            fmt_requests.append(_cell_format_request(
                sheet_id, r, c, c + 1, num_fmt='#,##0'))

    # 헤더 행 (Row 5)
    fmt_requests.append(_row_format_request(
        sheet_id, 5, bg='#333333', bold=True, font_color='#FFFFFF',
        h_align='CENTER', borders=True))

    # 데이터/소계/총계 행
    for ri, rtype, tier in row_meta:
        if rtype == 'data':
            fmt_requests.append(_row_format_request(
                sheet_id, ri, bg=TIER_COLORS[tier]['bg'], borders=True))
            # 금액 열 D~H (index 3~7)
            for c in range(3, 8):
                fmt_requests.append(_cell_format_request(
                    sheet_id, ri, c, c + 1, num_fmt='#,##0'))

        elif rtype == 'subtotal':
            fmt_requests.append(_row_format_request(
                sheet_id, ri, bg=TIER_COLORS[tier]['header'],
                bold=True, font_color='#FFFFFF', borders=True))
            for c in range(3, 8):
                fmt_requests.append(_cell_format_request(
                    sheet_id, ri, c, c + 1, num_fmt='#,##0'))

        elif rtype == 'total':
            fmt_requests.append(_row_format_request(
                sheet_id, ri, bg='#000000', bold=True,
                font_color='#FFFFFF', font_size=11, borders=True))
            for c in range(3, 8):
                fmt_requests.append(_cell_format_request(
                    sheet_id, ri, c, c + 1, num_fmt='#,##0'))

    # 고정 행 (헤더까지)
    fmt_requests.append({
        'updateSheetProperties': {
            'properties': {
                'sheetId': sheet_id,
                'gridProperties': {'frozenRowCount': 6},
            },
            'fields': 'gridProperties.frozenRowCount',
        }
    })

    # 서식 일괄 적용
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={'requests': fmt_requests}
    ).execute()

    print("서식 적용 완료 (색상, 테두리, 폰트, 열너비, 고정행)")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='새 예산안 시트 생성')
    parser.add_argument('--apply', action='store_true', help='Google Sheets에 직접 적용')
    args = parser.parse_args()

    if args.apply:
        create_google_sheet()
    else:
        path = create_xlsx_preview()
        print(f"\n미리보기 파일을 확인 후, --apply 옵션으로 Google Sheets에 적용하세요.")
