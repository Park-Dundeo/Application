from __future__ import annotations

from pathlib import Path
from datetime import datetime, date, time
import csv
import openpyxl
from app.config import AppConfig


STANDARD_HEADERS = [
    "날짜",
    "시간",
    "타입",
    "대분류",
    "소분류",
    "내용",
    "금액",
    "화폐",
    "결제수단",
    "메모",
    "상세",
    "원본파일",
    "원본행ID",
]


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def normalize_latest(cfg: AppConfig, unzip_dir: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = cfg.staging_dir / f"normalized_{stamp}.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # pick first xlsx file in unzip_dir
    xlsx_files = sorted(p for p in unzip_dir.iterdir() if p.suffix.lower() == ".xlsx")
    if not xlsx_files:
        raise RuntimeError(f"No xlsx found in {unzip_dir}")

    source_path = xlsx_files[0]
    wb = openpyxl.load_workbook(source_path, data_only=True)
    if "가계부 내역" not in wb.sheetnames:
        raise RuntimeError("Sheet '가계부 내역' not found in export")

    ws = wb["가계부 내역"]

    # Map headers from first row
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(v).strip(): idx + 1 for idx, v in enumerate(header_row) if v}

    def get_cell(row_idx: int, header_name: str) -> str:
        col = header_map.get(header_name)
        if not col:
            return ""
        return _format_cell(ws.cell(row=row_idx, column=col).value)

    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=STANDARD_HEADERS)
        writer.writeheader()

        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(row=r, column=header_map.get("날짜", 1)).value
            if date_val is None:
                continue

            row = {
                "날짜": get_cell(r, "날짜"),
                "시간": get_cell(r, "시간"),
                "타입": get_cell(r, "타입"),
                "대분류": get_cell(r, "대분류"),
                "소분류": get_cell(r, "소분류"),
                "내용": get_cell(r, "내용"),
                "금액": get_cell(r, "금액"),
                "화폐": get_cell(r, "화폐"),
                "결제수단": get_cell(r, "결제수단"),
                "메모": get_cell(r, "메모"),
                "상세": "",
                "원본파일": source_path.name,
                "원본행ID": str(r),
            }
            writer.writerow(row)

    return out_path
