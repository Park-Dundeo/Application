"""
청지기 재정 예산 파이프라인 (Config-Driven)

budget_config.yaml을 읽어서 Google Sheets 예산안 시트를 생성/갱신합니다.

사용법:
  python -m app.pipeline.budget deploy          # 시트 생성
  python -m app.pipeline.budget deploy --force   # 기존 시트 삭제 후 재생성
  python -m app.pipeline.budget status           # 예산 현황 요약
  python -m app.pipeline.budget validate         # config 무결성 검증
  python -m app.pipeline.budget preview          # XLSX 미리보기 생성
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml

from app.config import AppConfig


# ── Config 로딩 ──────────────────────────────────────────


@dataclass
class BudgetItem:
    key: str
    monthly: int
    note: str = ""
    item_type: str = "regular"  # regular | irregular
    annual_from_bonus: int = 0


@dataclass
class Project:
    id: str
    name: str
    goal: str = ""
    note: str = ""
    items: list[BudgetItem] = field(default_factory=list)


@dataclass
class Tier:
    id: str
    name: str
    priority: int
    philosophy: str = ""
    color_bg: str = "#FFFFFF"
    color_header: str = "#333333"
    projects: list[Project] = field(default_factory=list)


@dataclass
class BudgetConfig:
    period: str
    period_start: str
    period_end: str
    monthly_base: int
    annual_bonus: int
    income_description: str
    tiers: list[Tier] = field(default_factory=list)
    bonus_allocation: dict[str, int] = field(default_factory=dict)

    @property
    def annual_income(self) -> int:
        return self.monthly_base * 12 + self.annual_bonus

    def all_items(self) -> list[tuple[Tier, Project, BudgetItem]]:
        """모든 항목을 (tier, project, item) 튜플 리스트로 반환"""
        result = []
        for tier in self.tiers:
            for project in tier.projects:
                for item in project.items:
                    result.append((tier, project, item))
        return result

    def total_monthly(self) -> int:
        return sum(item.monthly for _, _, item in self.all_items())

    def total_annual(self) -> int:
        return sum(item.monthly * 12 for _, _, item in self.all_items())

    def tier_monthly(self, tier_id: str) -> int:
        for tier in self.tiers:
            if tier.id == tier_id:
                return sum(
                    item.monthly
                    for proj in tier.projects
                    for item in proj.items
                )
        return 0


def load_budget_config(config_path: Path | None = None) -> BudgetConfig:
    """budget_config.yaml 파일을 파싱하여 BudgetConfig 객체로 반환"""
    if config_path is None:
        config_path = Path("data/budget_config.yaml")

    with open(config_path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f)

    tiers = []
    for t in raw.get("tiers", []):
        projects = []
        for p in t.get("projects", []):
            items = []
            for i in p.get("items", []):
                items.append(BudgetItem(
                    key=i["key"],
                    monthly=i.get("monthly", 0),
                    note=i.get("note", ""),
                    item_type=i.get("type", "regular"),
                    annual_from_bonus=i.get("annual_from_bonus", 0),
                ))
            projects.append(Project(
                id=p["id"],
                name=p["name"],
                goal=p.get("goal", ""),
                note=p.get("note", ""),
                items=items,
            ))
        color = t.get("color", {})
        tiers.append(Tier(
            id=t["id"],
            name=t["name"],
            priority=t.get("priority", 99),
            philosophy=t.get("philosophy", ""),
            color_bg=color.get("bg", "#FFFFFF"),
            color_header=color.get("header", "#333333"),
            projects=projects,
        ))

    income = raw.get("income", {})
    return BudgetConfig(
        period=raw.get("period", ""),
        period_start=raw.get("period_start", ""),
        period_end=raw.get("period_end", ""),
        monthly_base=income.get("monthly_base", 0),
        annual_bonus=income.get("annual_bonus", 0),
        income_description=income.get("description", ""),
        tiers=tiers,
        bonus_allocation=raw.get("bonus_allocation", {}),
    )


# ── Google Sheets 서식 헬퍼 ──────────────────────────────

NUM_COLS = 8  # A~H


def _hex_to_rgb(hex_color: str) -> dict[str, float]:
    h = hex_color.lstrip("#")
    return {
        "red": int(h[0:2], 16) / 255,
        "green": int(h[2:4], 16) / 255,
        "blue": int(h[4:6], 16) / 255,
    }


def _cell_fmt(sheet_id: int, row: int, col: int,
              col_end: int | None = None, **kwargs: Any) -> dict:
    if col_end is None:
        col_end = col + 1
    fields = []
    cell_format: dict[str, Any] = {}

    if bg := kwargs.get("bg"):
        cell_format["backgroundColor"] = _hex_to_rgb(bg)
        fields.append("userEnteredFormat.backgroundColor")

    tf: dict[str, Any] = {}
    if kwargs.get("bold") is not None:
        tf["bold"] = kwargs["bold"]
    if fc := kwargs.get("font_color"):
        tf["foregroundColor"] = _hex_to_rgb(fc)
    if fs := kwargs.get("font_size"):
        tf["fontSize"] = fs
    if tf:
        cell_format["textFormat"] = tf
        fields.append("userEnteredFormat.textFormat")

    if ha := kwargs.get("h_align"):
        cell_format["horizontalAlignment"] = ha
        fields.append("userEnteredFormat.horizontalAlignment")

    if va := kwargs.get("v_align"):
        cell_format["verticalAlignment"] = va
        fields.append("userEnteredFormat.verticalAlignment")

    if nf := kwargs.get("num_fmt"):
        cell_format["numberFormat"] = {"type": "NUMBER", "pattern": nf}
        fields.append("userEnteredFormat.numberFormat")

    if kwargs.get("borders"):
        bs = {"style": "SOLID", "color": _hex_to_rgb("#CCCCCC")}
        cell_format["borders"] = {
            "top": bs, "bottom": bs, "left": bs, "right": bs
        }
        fields.append("userEnteredFormat.borders")

    if kwargs.get("wrap"):
        cell_format["wrapStrategy"] = "WRAP"
        fields.append("userEnteredFormat.wrapStrategy")

    return {
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": row,
                "endRowIndex": row + 1,
                "startColumnIndex": col,
                "endColumnIndex": col_end,
            },
            "cell": {"userEnteredFormat": cell_format},
            "fields": ",".join(fields),
        }
    }


def _row_fmt(sheet_id: int, row: int, col_end: int = NUM_COLS,
             **kwargs: Any) -> dict:
    return _cell_fmt(sheet_id, row, 0, col_end, **kwargs)


def _merge(sheet_id: int, row: int, col_start: int = 0,
           col_end: int = NUM_COLS) -> dict:
    """병합 요청 생성"""
    return {
        "mergeCells": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": row,
                "endRowIndex": row + 1,
                "startColumnIndex": col_start,
                "endColumnIndex": col_end,
            },
            "mergeType": "MERGE_ALL",
        }
    }


# ── 시트 생성 (deploy) ──────────────────────────────────


def _build_layout(cfg: BudgetConfig) -> dict:
    """새 비주얼 레이아웃 데이터를 구성하여 반환.

    Returns dict with keys:
        values: list[list]  — 시트에 쓸 2D 배열
        row_meta: list[tuple[int, str, Tier|None, dict|None]]
        merges: list[int]   — 풀너비 병합할 0-indexed row 목록
        tier_summary_rows: list[int]
        detail_start_row: int
        total_row: int
    """
    ledger_name = "가계부 내역"
    ps = cfg.period_start
    pe = cfg.period_end
    ps_year, ps_month, ps_day = ps.split("-")
    pe_year, pe_month, pe_day = pe.split("-")

    values: list[list] = []
    row_meta: list[tuple] = []
    merges: list[int] = []

    # ── Zone 1: 헤더 ──
    # Row 0: 제목
    values.append([f"{cfg.period} 청지기 재정 예산안"])
    row_meta.append((0, "title", None, None))
    merges.append(0)

    # Row 1: 부제
    total_m = cfg.total_monthly()
    total_a = cfg.total_annual()
    ratio = total_a / cfg.annual_income * 100 if cfg.annual_income else 0
    values.append([
        f"월 수입 {cfg.monthly_base:,} · "
        f"연간 {cfg.annual_income:,} · "
        f"배분율 {ratio:.1f}%"
    ])
    row_meta.append((1, "subtitle", None, None))
    merges.append(1)

    # ── Zone 2: 5층 요약 카드 ──
    # Row 2: 빈 행
    values.append([])
    row_meta.append((2, "blank", None, None))

    # Row 3: 섹션 라벨
    values.append(["── 층별 요약 ──"])
    row_meta.append((3, "section_label", None, None))
    merges.append(3)

    # Row 4~8: 5층 요약 (tier마다 1행)
    tier_summary_rows: list[int] = []
    for tier in cfg.tiers:
        ri = len(values)
        t_monthly = cfg.tier_monthly(tier.id)
        t_annual = t_monthly * 12
        # 상여금 포함 연 총액 계산
        bonus_total = sum(
            item.annual_from_bonus
            for proj in tier.projects
            for item in proj.items
        )
        t_annual += bonus_total
        pct = t_annual / cfg.annual_income if cfg.annual_income else 0
        bar_len = max(1, round(pct * 40))
        values.append([
            "",
            f"{tier.priority}층 {tier.name}",
            t_monthly,
            t_annual,
            "", "",
            pct,
            "█" * bar_len,
        ])
        row_meta.append((ri, "tier_summary", tier, None))
        tier_summary_rows.append(ri)

    # ── Zone 3: 총계 바 ──
    # 빈 행
    ri_blank = len(values)
    values.append([])
    row_meta.append((ri_blank, "blank", None, None))

    # 총계 행
    total_summary_ri = len(values)
    values.append([
        "", "총계",
        total_m, total_a,
        "", "",
        total_a / cfg.annual_income if cfg.annual_income else 0,
        "",
    ])
    row_meta.append((total_summary_ri, "total_summary", None, None))

    # ── Zone 4: 구분선 (빈 행) ──
    ri_sep = len(values)
    values.append([])
    row_meta.append((ri_sep, "blank", None, None))

    # ── Zone 5: 층별 상세 ──
    detail_start_row = len(values)
    subtotal_indices: list[int] = []

    for tier in cfg.tiers:
        # 층 배너 행
        banner_ri = len(values)
        values.append([f"{tier.priority}층 {tier.name} — {tier.philosophy}"])
        row_meta.append((banner_ri, "tier_banner", tier, None))
        merges.append(banner_ri)

        # 칼럼 헤더 행
        col_hdr_ri = len(values)
        values.append([
            "", "프로젝트·항목", "월 예산", "연 예산",
            "실적", "잔여", "소진율", "비고",
        ])
        row_meta.append((col_hdr_ri, "col_header", tier, None))

        # 프로젝트 그룹
        tier_data_start = len(values)  # 층 내 데이터 시작
        for proj in tier.projects:
            proj_data_start = len(values)
            first_in_proj = True
            for item in proj.items:
                ri = len(values)
                gs = ri + 1
                annual = item.monthly * 12
                note_parts = []
                if item.note:
                    note_parts.append(item.note)
                if item.annual_from_bonus:
                    note_parts.append(f"상여금 {item.annual_from_bonus:,}")
                    annual += item.annual_from_bonus

                note_str = " / ".join(note_parts) if note_parts else ""
                sumif = (
                    f"=-SUMIFS('{ledger_name}'!$G:$G,"
                    f"'{ledger_name}'!$K:$K,B{gs},"
                    f"'{ledger_name}'!$A:$A,\">=\"&DATE({ps_year},{ps_month},{ps_day}),"
                    f"'{ledger_name}'!$A:$A,\"<=\"&DATE({pe_year},{pe_month},{pe_day}))"
                )
                pct_formula = f'=IF(D{gs}=0,"",E{gs}/D{gs})'
                bar_formula = (
                    f'=IF(G{gs}="","",REPT("█",MIN(ROUND(G{gs}*20),20)))'
                    + (f'&" · {note_str}"' if note_str else "")
                )

                values.append([
                    proj.name if first_in_proj else "",
                    item.key,
                    item.monthly,
                    annual,
                    sumif,
                    f"=D{gs}-E{gs}",
                    pct_formula,
                    bar_formula,
                ])
                row_meta.append((ri, "data", tier, {"project": proj}))
                first_in_proj = False

            # 프로젝트 소계 (항목 2개 이상일 때만)
            if len(proj.items) > 1:
                ri = len(values)
                gs = ri + 1
                ps_start = proj_data_start + 1
                ps_end = gs - 1
                values.append([
                    "",
                    f"── {proj.name} 소계 ──",
                    f"=SUM(C{ps_start}:C{ps_end})",
                    f"=SUM(D{ps_start}:D{ps_end})",
                    f"=SUM(E{ps_start}:E{ps_end})",
                    f"=SUM(F{ps_start}:F{ps_end})",
                    f'=IF(SUM(D{ps_start}:D{ps_end})=0,"",SUM(E{ps_start}:E{ps_end})/SUM(D{ps_start}:D{ps_end}))',
                    "",
                ])
                row_meta.append((ri, "proj_subtotal", tier, {"project": proj}))

        # 층 소계 행
        tier_sub_ri = len(values)
        gs = tier_sub_ri + 1
        ts_start = tier_data_start + 1
        ts_end = gs - 1
        # SUM only data rows (skip proj_subtotal rows to avoid double-count)
        # Use individual cell refs for data rows only
        data_rows_in_tier = [
            r + 1  # 1-indexed
            for r, rtype, t, _ in row_meta
            if rtype == "data" and t and t.id == tier.id
        ]
        c_refs = "+".join(f"C{r}" for r in data_rows_in_tier)
        d_refs = "+".join(f"D{r}" for r in data_rows_in_tier)
        e_refs = "+".join(f"E{r}" for r in data_rows_in_tier)
        f_refs = "+".join(f"F{r}" for r in data_rows_in_tier)

        values.append([
            "",
            f"{tier.name} 소계",
            f"={c_refs}" if c_refs else 0,
            f"={d_refs}" if d_refs else 0,
            f"={e_refs}" if e_refs else 0,
            f"={f_refs}" if f_refs else 0,
            f'=IF(({d_refs})=0,"",({e_refs})/({d_refs}))' if d_refs else "",
            "",
        ])
        row_meta.append((tier_sub_ri, "tier_subtotal", tier, None))
        subtotal_indices.append(tier_sub_ri)

        # 빈 행 (층 간 간격)
        blank_ri = len(values)
        values.append([])
        row_meta.append((blank_ri, "blank", None, None))

    # ── 총계 행 (상세 영역) ──
    total_ri = len(values)
    gs_total = total_ri + 1

    def _sub_refs(col: str) -> str:
        return "+".join(f"{col}{si + 1}" for si in subtotal_indices)

    c_sub = _sub_refs("C")
    d_sub = _sub_refs("D")
    e_sub = _sub_refs("E")
    f_sub = _sub_refs("F")

    values.append([
        "", "총계",
        f"={c_sub}", f"={d_sub}",
        f"={e_sub}", f"={f_sub}",
        f'=IF(({d_sub})=0,"",({e_sub})/({d_sub}))',
        f'=IF(G{gs_total}="","",REPT("█",MIN(ROUND(G{gs_total}*20),20)))',
    ])
    row_meta.append((total_ri, "total", None, None))

    # 총계 요약 Zone 3의 실적/잔여를 수식으로 업데이트
    ts_gs = total_summary_ri + 1
    t_gs = gs_total
    values[total_summary_ri][4] = f"=E{t_gs}"  # 실적
    values[total_summary_ri][5] = f"=F{t_gs}"  # 잔여

    return {
        "values": values,
        "row_meta": row_meta,
        "merges": merges,
        "tier_summary_rows": tier_summary_rows,
        "detail_start_row": detail_start_row,
        "total_row": total_ri,
        "subtotal_indices": subtotal_indices,
    }


def deploy(cfg: BudgetConfig, spreadsheet_id: str, force: bool = False) -> None:
    """config 기반으로 Google Sheets 예산안 시트 생성 (비주얼 레이아웃)"""
    from googleapiclient.discovery import build
    from app.adapters.google_auth import get_credentials

    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = get_credentials(SCOPES)
    service = build("sheets", "v4", credentials=creds)

    sheet_title = f"{cfg.period} 예산안"

    # 기존 시트 확인
    resp = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title))"
    ).execute()

    existing_id = None
    for sheet in resp.get("sheets", []):
        if sheet["properties"]["title"] == sheet_title:
            existing_id = sheet["properties"]["sheetId"]
            break

    if existing_id is not None:
        if force:
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": [{"deleteSheet": {"sheetId": existing_id}}]}
            ).execute()
            print(f"기존 시트 삭제: {sheet_title}")
        else:
            print(f"ERROR: '{sheet_title}' 시트가 이미 존재합니다.")
            print("  --force 옵션으로 재생성하거나, 시트 이름을 변경하세요.")
            sys.exit(1)

    # 레이아웃 구성
    layout = _build_layout(cfg)
    values = layout["values"]
    row_meta = layout["row_meta"]
    merge_rows = layout["merges"]
    total_row = layout["total_row"]

    # 시트 생성
    result = service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{
            "addSheet": {
                "properties": {
                    "title": sheet_title,
                    "gridProperties": {
                        "rowCount": max(len(values) + 5, 80),
                        "columnCount": NUM_COLS,
                    },
                }
            }
        }]}
    ).execute()
    sheet_id = result["replies"][0]["addSheet"]["properties"]["sheetId"]
    print(f"시트 생성: {sheet_title} (ID: {sheet_id})")

    # ── 데이터 입력 ──
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_title}'!A1",
        valueInputOption="USER_ENTERED",
        body={"values": values},
    ).execute()
    print(f"데이터 입력: {len(values)}행")

    # ── 서식 적용 ──
    fmt: list[dict] = []

    # 열 너비 (A~H)
    col_widths = [50, 180, 100, 120, 120, 100, 80, 200]
    for i, px in enumerate(col_widths):
        fmt.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id, "dimension": "COLUMNS",
                    "startIndex": i, "endIndex": i + 1,
                },
                "properties": {"pixelSize": px},
                "fields": "pixelSize",
            }
        })

    # 병합
    for mr in merge_rows:
        fmt.append(_merge(sheet_id, mr))

    # 행별 서식
    for ri, rtype, tier_obj, extra in row_meta:
        if rtype == "title":
            fmt.append(_row_fmt(sheet_id, ri, bold=True, font_size=16,
                                bg="#263238", font_color="#FFFFFF"))
        elif rtype == "subtitle":
            fmt.append(_row_fmt(sheet_id, ri, font_size=11,
                                font_color="#757575", bg="#ECEFF1"))
        elif rtype == "section_label":
            fmt.append(_row_fmt(sheet_id, ri, font_size=10,
                                font_color="#9E9E9E", h_align="CENTER"))
        elif rtype == "tier_summary" and tier_obj:
            fmt.append(_row_fmt(sheet_id, ri, bg=tier_obj.color_bg, borders=True))
            # B열: bold
            fmt.append(_cell_fmt(sheet_id, ri, 1, 2, bold=True,
                                 font_color=tier_obj.color_header))
            # C,D열: 금액 서식
            for c in (2, 3):
                fmt.append(_cell_fmt(sheet_id, ri, c, c + 1, num_fmt="#,##0"))
            # G열: 비중 퍼센트
            fmt.append(_cell_fmt(sheet_id, ri, 6, 7, num_fmt="0.0%"))
            # H열: 바 — 층 고유색
            fmt.append(_cell_fmt(sheet_id, ri, 7, 8,
                                 font_color=tier_obj.color_header, font_size=9))
        elif rtype == "total_summary":
            fmt.append(_row_fmt(sheet_id, ri, bg="#000000", bold=True,
                                font_color="#FFFFFF", font_size=11, borders=True))
            for c in (2, 3, 4, 5):
                fmt.append(_cell_fmt(sheet_id, ri, c, c + 1,
                                     num_fmt="#,##0", bold=True,
                                     font_color="#FFFFFF"))
            fmt.append(_cell_fmt(sheet_id, ri, 6, 7, num_fmt="0.0%",
                                 bold=True, font_color="#FFFFFF"))
        elif rtype == "tier_banner" and tier_obj:
            fmt.append(_row_fmt(sheet_id, ri, bg=tier_obj.color_header,
                                bold=True, font_color="#FFFFFF", font_size=12))
        elif rtype == "col_header":
            fmt.append(_row_fmt(sheet_id, ri, bg="#F5F5F5",
                                font_size=9, h_align="CENTER",
                                font_color="#757575"))
        elif rtype == "data" and tier_obj:
            fmt.append(_row_fmt(sheet_id, ri, bg=tier_obj.color_bg, borders=True))
            # A열: 프로젝트명 — bold
            fmt.append(_cell_fmt(sheet_id, ri, 0, 1, bold=True))
            # C,D,E,F열: 금액 서식
            for c in (2, 3, 4, 5):
                fmt.append(_cell_fmt(sheet_id, ri, c, c + 1, num_fmt="#,##0"))
            # G열: 소진율 퍼센트
            fmt.append(_cell_fmt(sheet_id, ri, 6, 7, num_fmt="0%"))
            # H열: 바 font size
            fmt.append(_cell_fmt(sheet_id, ri, 7, 8, font_size=9, wrap=True))
        elif rtype == "proj_subtotal" and tier_obj:
            # 중간톤 — 층 bg보다 약간 진한 느낌 (헤더색에 투명도)
            fmt.append(_row_fmt(sheet_id, ri, bg=tier_obj.color_bg, borders=True,
                                bold=True, font_color=tier_obj.color_header))
            for c in (2, 3, 4, 5):
                fmt.append(_cell_fmt(sheet_id, ri, c, c + 1, num_fmt="#,##0",
                                     bold=True, font_color=tier_obj.color_header))
            fmt.append(_cell_fmt(sheet_id, ri, 6, 7, num_fmt="0%",
                                 bold=True, font_color=tier_obj.color_header))
        elif rtype == "tier_subtotal" and tier_obj:
            fmt.append(_row_fmt(sheet_id, ri, bg=tier_obj.color_header,
                                bold=True, font_color="#FFFFFF", borders=True))
            for c in (2, 3, 4, 5):
                fmt.append(_cell_fmt(sheet_id, ri, c, c + 1, num_fmt="#,##0",
                                     bold=True, font_color="#FFFFFF"))
            fmt.append(_cell_fmt(sheet_id, ri, 6, 7, num_fmt="0%",
                                 bold=True, font_color="#FFFFFF"))
        elif rtype == "total":
            fmt.append(_row_fmt(sheet_id, ri, bg="#000000", bold=True,
                                font_color="#FFFFFF", font_size=11, borders=True))
            for c in (2, 3, 4, 5):
                fmt.append(_cell_fmt(sheet_id, ri, c, c + 1, num_fmt="#,##0",
                                     bold=True, font_color="#FFFFFF"))
            fmt.append(_cell_fmt(sheet_id, ri, 6, 7, num_fmt="0%",
                                 bold=True, font_color="#FFFFFF"))

    # 고정 행 (제목 + 부제)
    fmt.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "gridProperties": {"frozenRowCount": 2},
            },
            "fields": "gridProperties.frozenRowCount",
        }
    })

    # ── 조건부 서식: G열(소진율) — 경과 시간 대비 동적 색상 ──
    # 경과율 = (TODAY()-기간시작)/(기간종료-기간시작)
    # 소진율을 경과율 대비로 비교하여 색상 결정
    detail_start = layout["detail_start_row"]
    data_end = total_row + 2
    g_range = {
        "sheetId": sheet_id,
        "startRowIndex": detail_start,
        "endRowIndex": data_end,
        "startColumnIndex": 6,  # G열
        "endColumnIndex": 7,
    }
    # 기간 시작/종료를 DATE 함수로 표현
    ps = cfg.period_start
    pe = cfg.period_end
    ps_y, ps_m, ps_d = ps.split("-")
    pe_y, pe_m, pe_d = pe.split("-")
    elapsed = (
        f"(TODAY()-DATE({ps_y},{ps_m},{ps_d}))"
        f"/(DATE({pe_y},{pe_m},{pe_d})-DATE({ps_y},{ps_m},{ps_d}))"
    )
    # G 셀 참조 (첫 데이터 행 기준, 조건부 서식은 상대 참조로 적용됨)
    g_ref = f"G{detail_start + 1}"
    cond_rules = [
        # 100%+ 또는 경과율의 1.5배 이상 → 빨강 (초과)
        {"addConditionalFormatRule": {"rule": {
            "ranges": [g_range],
            "booleanRule": {
                "condition": {
                    "type": "CUSTOM_FORMULA",
                    "values": [{"userEnteredValue":
                        f"=OR({g_ref}>=1, {g_ref}>={elapsed}*1.5)"}],
                },
                "format": {
                    "backgroundColor": _hex_to_rgb("#FFEBEE"),
                    "textFormat": {"foregroundColor": _hex_to_rgb("#C62828")},
                }
            }
        }, "index": 0}},
        # 경과율의 1.2~1.5배 → 주황 (주의)
        {"addConditionalFormatRule": {"rule": {
            "ranges": [g_range],
            "booleanRule": {
                "condition": {
                    "type": "CUSTOM_FORMULA",
                    "values": [{"userEnteredValue":
                        f"={g_ref}>={elapsed}*1.2"}],
                },
                "format": {
                    "backgroundColor": _hex_to_rgb("#FFF3E0"),
                    "textFormat": {"foregroundColor": _hex_to_rgb("#E65100")},
                }
            }
        }, "index": 1}},
        # 경과율의 0.8~1.2배 → 녹색 (양호)
        {"addConditionalFormatRule": {"rule": {
            "ranges": [g_range],
            "booleanRule": {
                "condition": {
                    "type": "CUSTOM_FORMULA",
                    "values": [{"userEnteredValue":
                        f"={g_ref}>={elapsed}*0.8"}],
                },
                "format": {
                    "backgroundColor": _hex_to_rgb("#E8F5E9"),
                    "textFormat": {"foregroundColor": _hex_to_rgb("#2E7D32")},
                }
            }
        }, "index": 2}},
        # 경과율의 0.8 미만 → 파랑 (여유)
        {"addConditionalFormatRule": {"rule": {
            "ranges": [g_range],
            "booleanRule": {
                "condition": {
                    "type": "CUSTOM_FORMULA",
                    "values": [{"userEnteredValue":
                        f"=AND({g_ref}>=0, {g_ref}<{elapsed}*0.8)"}],
                },
                "format": {
                    "backgroundColor": _hex_to_rgb("#E3F2FD"),
                    "textFormat": {"foregroundColor": _hex_to_rgb("#1565C0")},
                }
            }
        }, "index": 3}},
    ]
    fmt.extend(cond_rules)

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": fmt},
    ).execute()
    print("서식 적용 완료")


# ── validate ──────────────────────────────────────────────


def validate(cfg: BudgetConfig) -> bool:
    """config 무결성 검증"""
    ok = True
    keys_seen: dict[str, str] = {}  # key → "tier/project"

    for tier in cfg.tiers:
        if not tier.id:
            print(f"ERROR: 층 id가 비어있음")
            ok = False
        for proj in tier.projects:
            if not proj.id:
                print(f"ERROR: {tier.name} 내 프로젝트 id가 비어있음")
                ok = False
            for item in proj.items:
                if not item.key:
                    print(f"ERROR: {tier.name}/{proj.name} 내 항목 key가 비어있음")
                    ok = False
                loc = f"{tier.name}/{proj.name}"
                if item.key in keys_seen:
                    print(f"WARN: K열 키 중복 '{item.key}'"
                          f" — {keys_seen[item.key]} ↔ {loc}")
                keys_seen[item.key] = loc
                if item.monthly < 0:
                    print(f"ERROR: {loc}/{item.key} 월예산이 음수: {item.monthly}")
                    ok = False

    total = cfg.total_annual()
    income = cfg.annual_income
    ratio = total / income * 100 if income else 0
    print(f"\n연간 수입:     {income:>12,}원")
    print(f"연간 지출예산: {total:>12,}원")
    print(f"비율:          {ratio:>11.1f}%")
    if total > income:
        print(f"WARN: 지출예산이 수입을 {total - income:,}원 초과")

    print(f"\n층별 요약:")
    for tier in cfg.tiers:
        m = cfg.tier_monthly(tier.id)
        a = m * 12
        print(f"  {tier.priority}층 {tier.name:12s}  월 {m:>10,}  연 {a:>12,}")

    print(f"\nK열 키 총 {len(keys_seen)}개")

    if ok:
        print("\n검증 통과")
    else:
        print("\n검증 실패")
    return ok


# ── status ────────────────────────────────────────────────


def status(cfg: BudgetConfig) -> None:
    """예산 현황 요약 (로컬, Sheets 연결 없음)"""
    print(f"=== {cfg.period} 청지기 재정 예산 ===\n")

    print(f"수입: 월 {cfg.monthly_base:,} + 상여 {cfg.annual_bonus:,}"
          f" = 연 {cfg.annual_income:,}\n")

    for tier in cfg.tiers:
        m = cfg.tier_monthly(tier.id)
        proj_count = len(tier.projects)
        item_count = sum(len(p.items) for p in tier.projects)
        print(f"[{tier.priority}층] {tier.name}")
        print(f"  철학: {tier.philosophy}")
        print(f"  월 {m:,} (프로젝트 {proj_count}개, 항목 {item_count}개)")
        for proj in tier.projects:
            pm = sum(i.monthly for i in proj.items)
            goal_str = f" — {proj.goal}" if proj.goal else ""
            print(f"    {proj.name}: 월 {pm:,}{goal_str}")
        print()

    total_m = cfg.total_monthly()
    total_a = cfg.total_annual()
    remain = cfg.annual_income - total_a
    print(f"총계: 월 {total_m:,} / 연 {total_a:,}")
    print(f"잔여: 연 {remain:,} ({'흑자' if remain >= 0 else '적자'})")


# ── preview (XLSX) ────────────────────────────────────────


def preview(cfg: BudgetConfig) -> str:
    """XLSX 미리보기 생성 (새 비주얼 레이아웃)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import CellIsRule
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cfg.period} 예산안"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 열 너비 (A~H)
    for col_letter, w in {"A": 8, "B": 24, "C": 14, "D": 16,
                           "E": 16, "F": 14, "G": 10, "H": 30}.items():
        ws.column_dimensions[col_letter].width = w

    # _build_layout에서 데이터 가져오기
    layout = _build_layout(cfg)
    values = layout["values"]
    row_meta = layout["row_meta"]
    merge_rows = layout["merges"]

    def _fill(hex_color: str) -> PatternFill:
        h = hex_color.lstrip("#")
        return PatternFill(start_color=h, end_color=h, fill_type="solid")

    def _write_row(ws_row: int, data: list, merged: bool = False,
                   **style_kwargs):
        """행 데이터를 쓰고 스타일 적용. merged=True면 첫 셀에만 값 기록."""
        font = style_kwargs.get("font")
        fill = style_kwargs.get("fill")
        alignment = style_kwargs.get("alignment")
        border = style_kwargs.get("border")
        num_fmt_cols = style_kwargs.get("num_fmt_cols", {})
        for ci, val in enumerate(data, 1):
            if merged and ci > 1:
                # 병합 셀은 값 설정 불가, 스타일만
                cell = ws.cell(row=ws_row, column=ci)
            else:
                cell = ws.cell(row=ws_row, column=ci, value=val)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            if ci in num_fmt_cols:
                cell.number_format = num_fmt_cols[ci]

    # 행별로 쓰기
    for ri, rtype, tier_obj, extra in row_meta:
        ws_row = ri + 1  # openpyxl은 1-indexed
        row_data = values[ri] if ri < len(values) else []
        # 빈 행을 NUM_COLS로 패딩
        while len(row_data) < NUM_COLS:
            row_data.append("")

        if rtype == "title":
            _write_row(ws_row, row_data, merged=True,
                       font=Font(bold=True, size=16, color="FFFFFF"),
                       fill=_fill("#263238"))
            ws.merge_cells(start_row=ws_row, start_column=1,
                           end_row=ws_row, end_column=NUM_COLS)

        elif rtype == "subtitle":
            _write_row(ws_row, row_data, merged=True,
                       font=Font(size=11, color="757575"),
                       fill=_fill("#ECEFF1"))
            ws.merge_cells(start_row=ws_row, start_column=1,
                           end_row=ws_row, end_column=NUM_COLS)

        elif rtype == "section_label":
            _write_row(ws_row, row_data, merged=True,
                       font=Font(size=10, color="9E9E9E"),
                       alignment=Alignment(horizontal="center"))
            ws.merge_cells(start_row=ws_row, start_column=1,
                           end_row=ws_row, end_column=NUM_COLS)

        elif rtype == "tier_summary" and tier_obj:
            bg = _fill(tier_obj.color_bg)
            hdr_hex = tier_obj.color_header.lstrip("#")
            _write_row(ws_row, row_data,
                       fill=bg, border=thin,
                       num_fmt_cols={3: "#,##0", 4: "#,##0", 7: "0.0%"})
            # B열: bold + 층 고유색
            cell_b = ws.cell(row=ws_row, column=2)
            cell_b.font = Font(bold=True, color=hdr_hex)
            # H열: 바 — 층 고유색
            cell_h = ws.cell(row=ws_row, column=8)
            cell_h.font = Font(size=9, color=hdr_hex)

        elif rtype == "total_summary":
            _write_row(ws_row, row_data,
                       font=Font(bold=True, size=11, color="FFFFFF"),
                       fill=_fill("#000000"), border=thin,
                       num_fmt_cols={3: "#,##0", 4: "#,##0",
                                    5: "#,##0", 6: "#,##0", 7: "0.0%"})

        elif rtype == "tier_banner" and tier_obj:
            _write_row(ws_row, row_data, merged=True,
                       font=Font(bold=True, size=12, color="FFFFFF"),
                       fill=_fill(tier_obj.color_header))
            ws.merge_cells(start_row=ws_row, start_column=1,
                           end_row=ws_row, end_column=NUM_COLS)

        elif rtype == "col_header":
            _write_row(ws_row, row_data,
                       font=Font(size=9, color="757575"),
                       fill=_fill("#F5F5F5"),
                       alignment=Alignment(horizontal="center"))

        elif rtype == "data" and tier_obj:
            bg = _fill(tier_obj.color_bg)
            _write_row(ws_row, row_data,
                       fill=bg, border=thin,
                       num_fmt_cols={3: "#,##0", 4: "#,##0",
                                    5: "#,##0", 6: "#,##0", 7: "0%"})
            # A열: 프로젝트명 bold
            cell_a = ws.cell(row=ws_row, column=1)
            cell_a.font = Font(bold=True)
            # H열: 바 작은 폰트
            cell_h = ws.cell(row=ws_row, column=8)
            cell_h.font = Font(size=9)

        elif rtype == "proj_subtotal" and tier_obj:
            hdr_hex = tier_obj.color_header.lstrip("#")
            bg = _fill(tier_obj.color_bg)
            _write_row(ws_row, row_data,
                       font=Font(bold=True, color=hdr_hex),
                       fill=bg, border=thin,
                       num_fmt_cols={3: "#,##0", 4: "#,##0",
                                    5: "#,##0", 6: "#,##0", 7: "0%"})

        elif rtype == "tier_subtotal" and tier_obj:
            _write_row(ws_row, row_data,
                       font=Font(bold=True, color="FFFFFF"),
                       fill=_fill(tier_obj.color_header), border=thin,
                       num_fmt_cols={3: "#,##0", 4: "#,##0",
                                    5: "#,##0", 6: "#,##0", 7: "0%"})

        elif rtype == "total":
            _write_row(ws_row, row_data,
                       font=Font(bold=True, size=11, color="FFFFFF"),
                       fill=_fill("#000000"), border=thin,
                       num_fmt_cols={3: "#,##0", 4: "#,##0",
                                    5: "#,##0", 6: "#,##0", 7: "0%"})

        elif rtype == "blank":
            pass  # 빈 행

    # ── 조건부 서식: G열(소진율) — 경과 시간 대비 동적 색상 ──
    # XLSX preview는 정적 수치 사용 (현재 경과율 기준)
    from datetime import date
    ps_dt = date.fromisoformat(cfg.period_start)
    pe_dt = date.fromisoformat(cfg.period_end)
    today = date.today()
    elapsed_ratio = max(0.01, (today - ps_dt).days / max(1, (pe_dt - ps_dt).days))

    detail_start = layout["detail_start_row"] + 1  # 1-indexed
    total_end = layout["total_row"] + 1  # 1-indexed
    g_range = f"G{detail_start}:G{total_end}"

    # 빨강: 100%+ 또는 경과율 1.5배 이상 (초과)
    ws.conditional_formatting.add(g_range, CellIsRule(
        operator="greaterThanOrEqual", formula=[str(min(1.0, elapsed_ratio * 1.5))],
        fill=PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
        font=Font(color="C62828")))
    # 주황: 경과율 1.2배 이상 (주의)
    ws.conditional_formatting.add(g_range, CellIsRule(
        operator="greaterThanOrEqual", formula=[str(elapsed_ratio * 1.2)],
        fill=PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        font=Font(color="E65100")))
    # 녹색: 경과율 0.8배 이상 (양호)
    ws.conditional_formatting.add(g_range, CellIsRule(
        operator="greaterThanOrEqual", formula=[str(elapsed_ratio * 0.8)],
        fill=PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        font=Font(color="2E7D32")))
    # 파랑: 경과율 0.8 미만 (여유)
    ws.conditional_formatting.add(g_range, CellIsRule(
        operator="greaterThanOrEqual", formula=["0"],
        fill=PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        font=Font(color="1565C0")))

    out_path = f"accountbook_analysis/budget_{cfg.period}.xlsx"
    wb.save(out_path)
    print(f"미리보기 생성: {out_path}")
    return out_path


# ── pipeline 통합용 (기존 인터페이스 유지) ────────────────


def refresh_budget_views(cfg: AppConfig) -> None:
    """기존 파이프라인 호환용. 추후 자동 갱신 로직 추가."""
    return


# ── CLI ───────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(
        description="청지기 재정 예산 관리",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
commands:
  deploy     Google Sheets에 예산안 시트 생성
  validate   config 무결성 검증
  status     예산 현황 요약 (로컬)
  preview    XLSX 미리보기 생성
        """,
    )
    parser.add_argument("command", choices=["deploy", "validate", "status", "preview"])
    parser.add_argument("--config", type=Path, default=Path("data/budget_config.yaml"),
                        help="budget config 경로")
    parser.add_argument("--force", action="store_true",
                        help="deploy 시 기존 시트 삭제 후 재생성")

    args = parser.parse_args()

    budget_cfg = load_budget_config(args.config)

    if args.command == "validate":
        return 0 if validate(budget_cfg) else 1

    elif args.command == "status":
        status(budget_cfg)
        return 0

    elif args.command == "preview":
        preview(budget_cfg)
        return 0

    elif args.command == "deploy":
        import os
        sid = os.environ.get("SPREADSHEET_ID", "")
        if not sid:
            print("ERROR: SPREADSHEET_ID 환경변수 필요")
            return 1
        deploy(budget_cfg, sid, force=args.force)
        return 0

    return 1


if __name__ == "__main__":
    sys.exit(main())
